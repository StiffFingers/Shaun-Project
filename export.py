"""Build Excel workbooks from journal entries."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("entry_date", "Date"),
    ("worker_name", "Worker"),
    ("project_name", "Project / Job Site"),
    ("weather", "Weather"),
    ("hours_worked", "Hours"),
    ("work_done", "Work Performed"),
    ("crew_notes", "Crew Notes"),
    ("materials_notes", "Materials"),
    ("issues_delays", "Issues / Delays"),
    ("safety_notes", "Safety Notes"),
    ("created_at", "Logged At"),
]


def _style_header(ws, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def _autosize(ws, min_width: int = 10, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            # Prefer first line length for wrapped cells
            lengths.append(max(len(line) for line in text.splitlines()) if text else 0)
        width = max(lengths) + 2 if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, width))


def build_excel(
    entries: list[dict[str, Any]],
    filters_summary: str = "",
) -> bytes:
    """Return .xlsx file bytes: All Entries + Summary by Worker + Summary by Project."""
    wb = Workbook()

    # --- Sheet 1: All Entries ---
    ws = wb.active
    ws.title = "All Entries"

    headers = [label for _, label in COLUMNS]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(headers))

    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    wrap = Alignment(vertical="top", wrap_text=True)

    # Sort oldest → newest for spreadsheet overview
    sorted_entries = sorted(
        entries,
        key=lambda e: (e.get("entry_date") or "", e.get("id") or 0),
    )

    for row_idx, entry in enumerate(sorted_entries, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            value = entry.get(key, "")
            if key == "hours_worked":
                try:
                    value = float(value or 0)
                except (TypeError, ValueError):
                    value = 0.0
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            cell.border = thin
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if key == "hours_worked":
                cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    _autosize(ws)
    # Wider text columns
    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        if key in ("work_done", "crew_notes", "materials_notes", "issues_delays", "safety_notes"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 40

    # --- Sheet 2: Summary by Worker ---
    ws_w = wb.create_sheet("By Worker")
    ws_w.append(["Worker", "Entries", "Total Hours", "Projects"])
    _style_header(ws_w, 4)

    by_worker: dict[str, dict[str, Any]] = {}
    for e in sorted_entries:
        name = e.get("worker_name") or "Unknown"
        bucket = by_worker.setdefault(name, {"count": 0, "hours": 0.0, "projects": set()})
        bucket["count"] += 1
        bucket["hours"] += float(e.get("hours_worked") or 0)
        bucket["projects"].add(e.get("project_name") or "")

    for name in sorted(by_worker):
        b = by_worker[name]
        ws_w.append([name, b["count"], round(b["hours"], 2), ", ".join(sorted(p for p in b["projects"] if p))])

    for row in ws_w.iter_rows(min_row=2, max_row=ws_w.max_row, max_col=4):
        for cell in row:
            cell.border = thin
            cell.alignment = wrap
        row[2].number_format = "0.00"
    _autosize(ws_w)
    ws_w.freeze_panes = "A2"

    # --- Sheet 3: Summary by Project ---
    ws_p = wb.create_sheet("By Project")
    ws_p.append(["Project / Job Site", "Entries", "Total Hours", "Workers"])
    _style_header(ws_p, 4)

    by_project: dict[str, dict[str, Any]] = {}
    for e in sorted_entries:
        name = e.get("project_name") or "Unknown"
        bucket = by_project.setdefault(name, {"count": 0, "hours": 0.0, "workers": set()})
        bucket["count"] += 1
        bucket["hours"] += float(e.get("hours_worked") or 0)
        bucket["workers"].add(e.get("worker_name") or "")

    for name in sorted(by_project):
        b = by_project[name]
        ws_p.append([name, b["count"], round(b["hours"], 2), ", ".join(sorted(w for w in b["workers"] if w))])

    for row in ws_p.iter_rows(min_row=2, max_row=ws_p.max_row, max_col=4):
        for cell in row:
            cell.border = thin
            cell.alignment = wrap
        row[2].number_format = "0.00"
    _autosize(ws_p)
    ws_p.freeze_panes = "A2"

    # --- Sheet 4: Cover / meta ---
    ws_m = wb.create_sheet("Export Info", 0)
    ws_m["A1"] = "Construction Work Journal — Export"
    ws_m["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_m["A3"] = "Generated"
    ws_m["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_m["A4"] = "Filters"
    ws_m["B4"] = filters_summary or "None (all entries)"
    ws_m["A5"] = "Total entries"
    ws_m["B5"] = len(sorted_entries)
    ws_m["A6"] = "Total hours"
    total_hours = sum(float(e.get("hours_worked") or 0) for e in sorted_entries)
    ws_m["B6"] = round(total_hours, 2)
    ws_m["A8"] = "Sheets"
    ws_m["B8"] = "All Entries — every log row"
    ws_m["B9"] = "By Worker — hours and entry counts per worker"
    ws_m["B10"] = "By Project — hours and entry counts per job site"
    ws_m.column_dimensions["A"].width = 16
    ws_m.column_dimensions["B"].width = 55

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
